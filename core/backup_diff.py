"""백업 비교 — 두 분기 백업의 차이 분석.

백업 폴더 안의 outstanding_members_*.xlsx 파일을 읽어
신규 진입 / 승급 / 강등 / 빠짐 / 변화 없음 으로 분류한다.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from config import BACKUPS_DIR, LEVEL_TEXT_MAP


@dataclass
class BackupSnapshot:
    folder: Path
    members: dict[str, dict] = field(default_factory=dict)

    @property
    def date_label(self) -> str:
        return self.folder.name

    def __len__(self) -> int:
        return len(self.members)


@dataclass
class BackupDiff:
    old: BackupSnapshot
    new: BackupSnapshot
    new_entries: list[dict] = field(default_factory=list)
    promoted: list[dict] = field(default_factory=list)
    demoted: list[dict] = field(default_factory=list)
    dropped: list[dict] = field(default_factory=list)
    unchanged: list[dict] = field(default_factory=list)

    @property
    def summary(self) -> str:
        return (
            f"신규 {len(self.new_entries)}명 / "
            f"승급 {len(self.promoted)}명 / "
            f"강등 {len(self.demoted)}명 / "
            f"빠짐 {len(self.dropped)}명 / "
            f"유지 {len(self.unchanged)}명"
        )


def load_snapshot(folder: Path) -> BackupSnapshot:
    """백업 폴더 안의 우수회원 xlsx 파일을 로드."""
    folder = Path(folder)
    xlsx_files = sorted(folder.glob("outstanding_members_*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(
            f"{folder} 안에 outstanding_members_*.xlsx 파일이 없습니다."
        )

    wb = load_workbook(str(xlsx_files[0]), read_only=True)
    ws = wb.active

    members: dict[str, dict] = {}
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return BackupSnapshot(folder=folder, members={})

    headers = [str(h or "").strip() for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    # 폴백 인덱스 (헤더가 깨지거나 다를 때)
    col_id = idx.get("아이디", 0)
    col_nick = idx.get("닉네임", 1)
    col_lv = idx.get("등급", 2)
    col_last = idx.get("마지막접속일", 3)
    col_join = idx.get("가입일", 4)

    for row in rows[1:]:
        if not row or row[col_id] is None:
            continue
        user_id = str(row[col_id]).strip()
        if not user_id:
            continue
        level_label = str(row[col_lv] or "").strip() if col_lv < len(row) else ""
        # 라벨에서 레벨 번호 추출 (LEVEL_TEXT_MAP 미스매치 시 0)
        level = LEVEL_TEXT_MAP.get(level_label, 0)
        members[user_id] = {
            "user_id": user_id,
            "nickname": str(row[col_nick] or "").strip() if col_nick < len(row) else "",
            "level": level,
            "level_label": level_label,
            "last_login": str(row[col_last] or "").strip() if col_last < len(row) else "",
            "join_date": str(row[col_join] or "").strip() if col_join < len(row) else "",
        }

    return BackupSnapshot(folder=folder, members=members)


def diff_backups(old: BackupSnapshot, new: BackupSnapshot) -> BackupDiff:
    diff = BackupDiff(old=old, new=new)
    old_ids = set(old.members.keys())
    new_ids = set(new.members.keys())

    for uid in sorted(new_ids - old_ids):
        diff.new_entries.append(new.members[uid])

    for uid in sorted(old_ids - new_ids):
        diff.dropped.append(old.members[uid])

    for uid in sorted(old_ids & new_ids):
        old_m = old.members[uid]
        new_m = new.members[uid]
        rec = {
            "user_id": uid,
            "nickname": new_m["nickname"],
            "old_level": old_m["level"],
            "new_level": new_m["level"],
            "old_level_label": old_m["level_label"],
            "new_level_label": new_m["level_label"],
        }
        if new_m["level"] > old_m["level"]:
            diff.promoted.append(rec)
        elif new_m["level"] < old_m["level"]:
            diff.demoted.append(rec)
        else:
            diff.unchanged.append(rec)

    return diff


def list_backup_folders(backups_dir: Path | None = None) -> list[Path]:
    """outstanding_members_*.xlsx 파일이 있는 백업 하위 폴더(오래된 순)."""
    base = Path(backups_dir or BACKUPS_DIR)
    if not base.exists():
        return []
    out: list[Path] = []
    for child in sorted(base.iterdir()):
        if child.is_dir() and any(child.glob("outstanding_members_*.xlsx")):
            out.append(child)
    return out


def write_diff_report(diff: BackupDiff, path: Path) -> None:
    lines: list[str] = []
    lines.append(f"백업 비교 리포트: {diff.old.date_label} → {diff.new.date_label}")
    lines.append(
        f"이전 {len(diff.old.members)}명, 이후 {len(diff.new.members)}명"
    )
    lines.append(diff.summary)
    lines.append("=" * 60)

    def _block(title: str, items: Iterable[dict], formatter) -> None:
        items = list(items)
        if not items:
            return
        lines.append(f"[{title} {len(items)}명]")
        for m in items:
            lines.append("  " + formatter(m))
        lines.append("")

    _block(
        "신규 진입",
        sorted(diff.new_entries, key=lambda x: -x["level"]),
        lambda m: f"+ {m['user_id']:<15} {m['nickname']:<15} ({m['level_label']})",
    )
    _block(
        "승급",
        diff.promoted,
        lambda m: f"↑ {m['user_id']:<15} {m['nickname']:<15} {m['old_level_label']} → {m['new_level_label']}",
    )
    _block(
        "강등",
        diff.demoted,
        lambda m: f"↓ {m['user_id']:<15} {m['nickname']:<15} {m['old_level_label']} → {m['new_level_label']}",
    )
    _block(
        "빠짐 (이전엔 있었으나 이번엔 없음)",
        sorted(diff.dropped, key=lambda x: -x["level"]),
        lambda m: f"- {m['user_id']:<15} {m['nickname']:<15} (이전 {m['level_label']})",
    )

    Path(path).write_text("\n".join(lines), encoding="utf-8")
