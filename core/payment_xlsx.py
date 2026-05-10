"""자료실 구독비 — XLSX 내보내기.

3개 워크시트:
    구독_매트릭스  — 회원 × 최근 12개월 (화면의 매트릭스를 그대로 옮김)
    활성_구독     — 오늘 기준 활성 회원 한 줄씩 (만료 임박순)
    거래내역      — 모든 토스 입금, 단가표 매칭 결과 표시

backups/payments_YYYY-MM-DD.xlsx 로 저장. v0.4 분기 백업과 같은 폴더 패턴.
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import BACKUPS_DIR
from core.payment_matrix import cell_state, month_range, status_summary
from core.payment_store import lookup_months

if TYPE_CHECKING:
    from core.models import Member
    from core.payment_store import PaymentStore


_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
_CENTER = Alignment(horizontal="center")
_MATRIX_MONTHS = 12


def default_xlsx_path(today: date | None = None) -> Path:
    d = today or date.today()
    return Path(BACKUPS_DIR) / f"payments_{d.isoformat()}.xlsx"


def write_payment_xlsx(
    path: Path | str,
    store: PaymentStore,
    members: list[Member],
    today: date | None = None,
) -> Path:
    """3개 시트 XLSX 작성. 반환값 = 저장 경로."""
    today = today or date.today()
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # 첫 시트(기본 활성 시트) 를 매트릭스로 사용
    ws_matrix = wb.active
    ws_matrix.title = "구독_매트릭스"

    _write_matrix_sheet(ws_matrix, store, members, today)
    _write_active_sheet(wb.create_sheet("활성_구독"), store, members, today)
    _write_transactions_sheet(wb.create_sheet("거래내역"), store)

    wb.save(str(out_path))
    return out_path


# ---------- 시트별 ----------

def _write_matrix_sheet(ws, store: PaymentStore, members: list[Member], today: date) -> None:
    months = month_range(today, _MATRIX_MONTHS)
    headers = ["아이디", "이름", "닉네임", "현재 상태"] + [h.label for h in months]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    rows = _build_member_rows(store, members)
    # 만료 임박순 (가까운 만료가 위)
    rows.sort(key=lambda item: max((s.period_to for s in item[1]), default=date.min))

    for member, subs in rows:
        cells = [
            member.user_id,
            member.name or "",
            member.nickname or "",
            status_summary(subs, today=today),
        ]
        cells.extend(cell_state(subs, h) for h in months)
        ws.append(cells)

    ws.freeze_panes = "E2"  # 회원 정보 4열 + 헤더 행 고정
    _auto_width(ws, len(headers), min_extra=4, max_width=28)


def _write_active_sheet(ws, store: PaymentStore, members: list[Member], today: date) -> None:
    headers = ["회원ID", "이름", "닉네임", "시작일", "만료일", "남은일수", "총구독개월"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    members_by_uid = {m.user_id: m for m in members}
    active: list[tuple[str, str, str, date, date, int, int]] = []
    for uid, all_subs in _subs_by_member(store).items():
        latest = max(all_subs, key=lambda s: s.period_to)
        if latest.period_to < today:
            continue  # 만료
        m = members_by_uid.get(uid)
        days_left = (latest.period_to - today).days
        total_months = sum(s.months for s in all_subs)
        # 가장 이른 시작일을 "시작일" 로 (체이닝 누적 시작점)
        first_start = min(s.period_from for s in all_subs)
        active.append((
            uid,
            (m.name if m else "") or "",
            (m.nickname if m else "") or "",
            first_start,
            latest.period_to,
            days_left,
            total_months,
        ))

    # 만료 임박순
    active.sort(key=lambda t: t[4])

    for row in active:
        ws.append([
            row[0],
            row[1],
            row[2],
            row[3].isoformat(),
            row[4].isoformat(),
            row[5],
            row[6],
        ])

    ws.freeze_panes = "A2"
    _auto_width(ws, len(headers), min_extra=4, max_width=24)


def _write_transactions_sheet(ws, store: PaymentStore) -> None:
    headers = ["거래일시", "입금자명", "금액", "분류", "거래기관", "메모", "원본파일"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    txns = sorted(store.all_transactions(), key=lambda t: t.txn_at, reverse=True)
    for t in txns:
        months = lookup_months(t.amount)
        category = f"{months}개월" if months is not None else "기타"
        ws.append([
            t.txn_at.strftime("%Y-%m-%d %H:%M"),
            t.payer_name,
            t.amount,
            category,
            t.bank,
            t.memo,
            t.source_file,
        ])

    ws.freeze_panes = "A2"
    _auto_width(ws, len(headers), min_extra=4, max_width=30)


# ---------- 헬퍼 ----------

def _build_member_rows(store: PaymentStore, members: list[Member]) -> list:
    """구독 이력 있는 회원만 (member, subs[]) 반환."""
    from core.models import Member
    members_by_uid = {m.user_id: m for m in members}
    by_uid = _subs_by_member(store)
    out: list = []
    for uid, subs in by_uid.items():
        m = members_by_uid.get(uid) or Member(user_id=uid, name="(목록에 없음)")
        out.append((m, sorted(subs, key=lambda s: s.period_from)))
    return out


def _subs_by_member(store: PaymentStore) -> dict:
    by_uid: dict = {}
    for s in store.all_subscriptions():
        by_uid.setdefault(s.member_user_id, []).append(s)
    return by_uid


def _style_header_row(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _auto_width(ws, ncols: int, *, min_extra: int = 4, max_width: int = 30) -> None:
    """컬럼 너비를 헤더 + 데이터 길이에 맞춰. backup_service.py 패턴."""
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col).value
            if val is None:
                continue
            length = len(str(val))
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(max_len + min_extra, max_width)
