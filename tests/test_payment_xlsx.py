"""payment_xlsx — 3시트 XLSX 내보내기 단위 테스트."""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from core.models import Member
from core.payment_store import PaymentStore, compute_period
from core.payment_xlsx import default_xlsx_path, write_payment_xlsx
from core.toss_xlsx import TossDeposit


@pytest.fixture
def store(tmp_path: Path) -> PaymentStore:
    return PaymentStore(tmp_path / "payments.db")


def _seed(store: PaymentStore) -> None:
    """홍길동 1구독(3000원=1개월) + 홍길동 1기타(50000원) 거래 시드."""
    deposits = [
        TossDeposit(
            txn_at=datetime(2026, 5, 8, 10, 0, 0),
            payer_name="홍길동",
            amount=3000,
            bank="토스뱅크",
            counterparty_account="",
            memo="",
            source_file="t.xlsx",
        ),
        TossDeposit(
            txn_at=datetime(2026, 5, 8, 11, 0, 0),
            payer_name="홍길동",
            amount=50000,  # 단가표 외 → 기타
            bank="토스뱅크",
            counterparty_account="",
            memo="",
            source_file="t.xlsx",
        ),
    ]
    store.import_deposits(deposits)
    txn = next(t for t in store.transactions_for_payer("홍길동") if t.amount == 3000)
    pf, pt = compute_period(date(2026, 5, 8), 1)
    store.add_subscription(
        member_user_id="hong",
        transaction_id=txn.id,
        months=1,
        period_from=pf,
        period_to=pt,
    )


def test_default_path_uses_backups_dir():
    p = default_xlsx_path(today=date(2026, 5, 9))
    assert p.name == "payments_2026-05-09.xlsx"
    assert "backups" in str(p).replace("\\", "/")


def test_write_xlsx_empty_store_creates_three_sheets(store, tmp_path):
    out = write_payment_xlsx(tmp_path / "p.xlsx", store, members=[])
    assert out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == ["구독_매트릭스", "활성_구독", "거래내역"]


def test_write_xlsx_with_seed_populates_sheets(store, tmp_path):
    _seed(store)
    members = [Member(user_id="hong", name="홍길동", nickname="홍이", level=5)]
    out = write_payment_xlsx(
        tmp_path / "p.xlsx", store, members=members, today=date(2026, 5, 9)
    )

    wb = load_workbook(out)

    # 매트릭스 시트 — 헤더 + 회원 1줄
    ws_m = wb["구독_매트릭스"]
    assert ws_m.cell(row=1, column=1).value == "아이디"
    assert ws_m.cell(row=2, column=1).value == "hong"
    assert ws_m.cell(row=2, column=2).value == "홍길동"

    # 활성 구독 시트 — 5/8~6/7 활성, 5/9 기준 남은일수 = 29
    ws_a = wb["활성_구독"]
    assert ws_a.cell(row=1, column=1).value == "회원ID"
    assert ws_a.cell(row=2, column=1).value == "hong"
    assert ws_a.cell(row=2, column=4).value == "2026-05-08"  # 시작일
    assert ws_a.cell(row=2, column=5).value == "2026-06-07"  # 만료일
    assert ws_a.cell(row=2, column=6).value == 29            # 남은일수

    # 거래내역 시트 — 2건 (시간 desc → 11시 행이 먼저)
    ws_t = wb["거래내역"]
    assert ws_t.cell(row=1, column=1).value == "거래일시"
    assert ws_t.cell(row=2, column=2).value == "홍길동"
    assert ws_t.cell(row=2, column=3).value == 50000      # 11시 거래(더 최근) 먼저
    assert ws_t.cell(row=2, column=4).value == "기타"
    assert ws_t.cell(row=3, column=3).value == 3000
    assert ws_t.cell(row=3, column=4).value == "1개월"


def test_xlsx_header_is_frozen(store, tmp_path):
    out = write_payment_xlsx(tmp_path / "p.xlsx", store, members=[])
    wb = load_workbook(out)
    # 매트릭스: 회원 정보 4열 + 헤더 1행 동결 → "E2"
    assert wb["구독_매트릭스"].freeze_panes == "E2"
    # 다른 두 시트: A2
    assert wb["활성_구독"].freeze_panes == "A2"
    assert wb["거래내역"].freeze_panes == "A2"


def test_xlsx_column_widths_positive(store, tmp_path):
    """헤더만 있어도 컬럼 너비가 음수·0 이 아닌지 sanity."""
    out = write_payment_xlsx(tmp_path / "p.xlsx", store, members=[])
    wb = load_workbook(out)
    for name in ("구독_매트릭스", "활성_구독", "거래내역"):
        ws = wb[name]
        for col_letter, dim in ws.column_dimensions.items():
            assert dim.width > 0, f"{name}!{col_letter} 너비 0 이하"
