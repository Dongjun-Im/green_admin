"""자료실 접속 로그 — XLSX 내보내기.

backups/nas_access_log_YYYY-MM-DD.xlsx. 한 시트 (접속로그) 에 항목별 한 줄.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import BACKUPS_DIR
from core.nas_log_service import ACTION_LABELS, enrich_with_members

if TYPE_CHECKING:
    from core.models import Member
    from core.nas_log_store import NasLogFilter, NasLogStore


_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
_CENTER = Alignment(horizontal="center")


def default_xlsx_path(today: date | None = None) -> Path:
    d = today or date.today()
    return Path(BACKUPS_DIR) / f"nas_access_log_{d.isoformat()}.xlsx"


def write_nas_log_xlsx(
    path: Path | str,
    store: "NasLogStore",
    members: list["Member"],
    *,
    flt: "NasLogFilter | None" = None,
    today: date | None = None,
) -> Path:
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    entries = store.entries(flt)
    rows = enrich_with_members(entries, members)

    wb = Workbook()
    ws = wb.active
    ws.title = "접속로그"
    headers = [
        "시간", "회원", "DSM 아이디", "IP", "프로토콜",
        "동작", "카테고리", "파일", "전체 경로",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER

    for i, r in enumerate(rows, start=2):
        e = r.entry
        ws.cell(row=i, column=1, value=e.logged_at.replace("T", " "))
        ws.cell(row=i, column=2, value=r.display_name)
        ws.cell(row=i, column=3, value=e.dsm_user_id)
        ws.cell(row=i, column=4, value=e.ip)
        ws.cell(row=i, column=5, value=e.protocol)
        ws.cell(row=i, column=6, value=ACTION_LABELS.get(e.action, e.action))
        ws.cell(row=i, column=7, value=e.category)
        ws.cell(row=i, column=8, value=e.file_name)
        ws.cell(row=i, column=9, value=e.file_path)

    widths = [19, 22, 16, 16, 12, 10, 18, 28, 48]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:I{len(rows) + 1}"

    wb.save(out_path)
    return out_path
