"""토스뱅크 거래내역 XLSX 파서.

비밀번호로 암호화된 토스 거래내역 .xlsx 를 메모리에서 복호화한 뒤
입금 거래만 추려서 TossDeposit 리스트로 반환한다.

실제 파일 구조 (data/samples/ 의 샘플로 확정):
    행 1     : "토스뱅크 거래내역" (제목)
    행 2     : 성명          / <이름>
    행 3     : 계좌번호       / <계좌>
    행 4     : 조회기간       / <기간>
    행 5~8   : 안내 / 공백
    행 9     : 헤더
                B 거래 일시
                C 적요               ← 입금자명 (실명)
                D 거래 유형          ← '입금' / '출금' / '이자입금' …
                E 거래 기관
                F 계좌번호
                G 거래 금액           ← 양수=입금, 음수=출금
                H 거래 후 잔액
                I 메모
    행 10~   : 거래 데이터
"""
from __future__ import annotations

import hashlib
import io
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import msoffcrypto
import openpyxl

# 시스템 거래(이자) 식별 키워드 — 제외 대상.
_SYSTEM_KEYWORDS = ("이자",)


@dataclass(frozen=True)
class TossDeposit:
    """토스 거래내역 1행 중 회원 입금 후보로 추린 데이터."""
    txn_at: datetime
    payer_name: str       # 적요 (입금자 실명)
    amount: int           # 원 단위 정수
    bank: str             # 거래 기관
    counterparty_account: str
    memo: str
    source_file: str      # 원본 파일명 (파일 경로 제외)

    @property
    def dedup_hash(self) -> str:
        """동일 거래 식별용 — 같은 일시·이름·금액이면 동일 거래로 간주."""
        key = f"{self.txn_at.isoformat()}|{self.payer_name}|{self.amount}"
        return hashlib.sha256(key.encode("utf-8")).hexdigest()


class TossPasswordError(Exception):
    """비밀번호가 틀렸거나 파일 형식이 비정상."""


def parse_toss_xlsx(path: Path | str, password: str) -> list[TossDeposit]:
    """암호화된 토스 거래내역 XLSX 를 복호화해 입금 행만 반환.

    Args:
        path: .xlsx 파일 경로
        password: 토스 거래내역 비밀번호

    Returns:
        입금 행만 추린 TossDeposit 리스트. 시스템 이자·출금은 제외.

    Raises:
        TossPasswordError: 비밀번호 오류 또는 파일 손상.
        FileNotFoundError: 파일 없음.
    """
    src = Path(path)
    if not src.exists():
        raise FileNotFoundError(str(src))

    buf = io.BytesIO()
    try:
        with src.open("rb") as f:
            of = msoffcrypto.OfficeFile(f)
            of.load_key(password=password)
            of.decrypt(buf)
    except Exception as e:
        # msoffcrypto 는 비밀번호 오류를 다양한 예외로 던짐 — 통일.
        raise TossPasswordError(f"복호화 실패: {e}") from e

    buf.seek(0)
    try:
        # read_only=True 는 토스 파일에서 시트 dimension 을 1×1 로 잘못 인식하는
        # 케이스가 있어 사용 안 함. 거래내역은 기껏해야 수백 행이라 메모리도 무리 없음.
        wb = openpyxl.load_workbook(buf, data_only=True)
    except Exception as e:
        raise TossPasswordError(f"엑셀 파싱 실패: {e}") from e

    ws = wb[wb.sheetnames[0]]
    deposits: list[TossDeposit] = []
    for row in ws.iter_rows(min_row=10, values_only=True):
        # 컬럼 인덱스: B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8 (A=0 은 비어있음).
        if not row or row[1] is None:
            continue
        txn_at_raw = row[1]
        payer = (row[2] or "").strip() if isinstance(row[2], str) else ""
        txn_type = (row[3] or "").strip() if isinstance(row[3], str) else ""
        bank = (row[4] or "").strip() if isinstance(row[4], str) else ""
        counter_acct = (row[5] or "").strip() if isinstance(row[5], str) else ""
        amount_raw = row[6]
        memo = (row[8] or "").strip() if isinstance(row[8], str) else ""

        # 시스템 거래(이자) 제외
        if any(kw in payer for kw in _SYSTEM_KEYWORDS) or any(
            kw in txn_type for kw in _SYSTEM_KEYWORDS
        ):
            continue
        # 입금만 (양수)
        if not isinstance(amount_raw, (int, float)) or amount_raw <= 0:
            continue

        txn_at = _parse_txn_datetime(txn_at_raw)
        if txn_at is None:
            continue

        deposits.append(TossDeposit(
            txn_at=txn_at,
            payer_name=payer,
            amount=int(amount_raw),
            bank=bank,
            counterparty_account=counter_acct,
            memo=memo,
            source_file=src.name,
        ))

    wb.close()
    return deposits


def _parse_txn_datetime(value) -> datetime | None:
    """토스의 '거래 일시' 셀은 datetime 또는 'YYYY.MM.DD HH:MM:SS' 문자열로 옴."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y.%m.%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y.%m.%d", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None
