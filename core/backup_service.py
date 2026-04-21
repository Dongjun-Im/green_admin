"""우수회원(7~8레벨) 백업 서비스: TXT + xlsx 출력."""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Callable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import BACKUPS_DIR, LEVEL_LABELS, OUTSTANDING_LEVELS
from core.crawler import MemberCrawler
from core.models import BackupResult, Member


ProgressCB = Callable[[int, int], None]


class BackupService:
    OUTSTANDING_LEVELS = set(OUTSTANDING_LEVELS)

    def __init__(
        self,
        crawler: MemberCrawler,
        backup_dir: Optional[Path] = None,
    ) -> None:
        self.crawler = crawler
        self.backup_dir = Path(backup_dir or BACKUPS_DIR)

    def run(self, progress_cb: Optional[ProgressCB] = None) -> BackupResult:
        all_members = self.crawler.fetch_all_members(progress_cb=progress_cb)
        outstanding = [m for m in all_members if m.level in self.OUTSTANDING_LEVELS]

        # 등급 desc, 마지막접속일 desc
        outstanding.sort(
            key=lambda m: (
                -m.level,
                -(m.last_login_date.toordinal() if m.last_login_date else 0),
            )
        )

        today = date.today()
        day_dir = self.backup_dir / today.isoformat()
        day_dir.mkdir(parents=True, exist_ok=True)

        txt_path = day_dir / f"outstanding_members_{today.isoformat()}.txt"
        xlsx_path = day_dir / f"outstanding_members_{today.isoformat()}.xlsx"

        self._write_txt(outstanding, txt_path)
        self._write_xlsx(outstanding, xlsx_path)

        breakdown: dict[int, int] = {}
        for m in outstanding:
            breakdown[m.level] = breakdown.get(m.level, 0) + 1

        return BackupResult(
            txt_path=txt_path,
            xlsx_path=xlsx_path,
            member_count=len(outstanding),
            level_breakdown=breakdown,
        )

    def _write_txt(self, members: list[Member], path: Path) -> None:
        today_str = date.today().isoformat()
        lines: list[str] = []
        lines.append(f"초록등대 우수회원 백업 ({today_str})")

        breakdown = {}
        for m in members:
            breakdown[m.level] = breakdown.get(m.level, 0) + 1
        summary = ", ".join(
            f"{LEVEL_LABELS.get(lv, str(lv))} {cnt}명"
            for lv, cnt in sorted(breakdown.items(), reverse=True)
        )
        lines.append(f"총 {len(members)}명" + (f" ({summary})" if summary else ""))
        lines.append("=" * 60)

        # 등급별 그룹
        for lv in sorted(self.OUTSTANDING_LEVELS, reverse=True):
            group = [m for m in members if m.level == lv]
            if not group:
                continue
            lines.append(f"[{LEVEL_LABELS.get(lv, str(lv))}]")
            for i, m in enumerate(group, start=1):
                last = m.last_login_date.isoformat() if m.last_login_date else "알수없음"
                join = m.join_date.isoformat() if m.join_date else "알수없음"
                lines.append(
                    f"{i:>3}. {m.user_id:<15} {m.nickname:<15} "
                    f"마지막접속 {last}   가입 {join}"
                )
            lines.append("")

        path.write_text("\n".join(lines), encoding="utf-8")

    def _write_xlsx(self, members: list[Member], path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "우수회원 백업"

        headers = ["아이디", "닉네임", "등급", "마지막접속일", "가입일"]
        ws.append(headers)

        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="DDEBF7")
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for m in members:
            ws.append([
                m.user_id,
                m.nickname,
                LEVEL_LABELS.get(m.level, str(m.level)),
                m.last_login_date.isoformat() if m.last_login_date else "",
                m.join_date.isoformat() if m.join_date else "",
            ])

        ws.freeze_panes = "A2"

        # 컬럼 너비 자동
        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            max_len = len(headers[col_idx - 1])
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[letter].width = min(max_len + 4, 40)

        wb.save(str(path))
